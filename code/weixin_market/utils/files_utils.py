# -*- coding: utf-8 -*-

'''
文件以及excel等相关的操作
'''
import os
import openpyxl
from openpyxl.workbook import Workbook
import requests
from io import BytesIO
import time
import json

'''
保存到本地目录, 最终存放的地址为 dir/file_name
file 文件对象
dir  本地的绝对文件路径
file_name 保存的名字, 不带路径的
return: True, dir/file_name 最终的存放地址
'''
def save_file(file, file_name, dir = ""):
    dest_file_name = dir + '/' + file_name

    try:
        # 创建目录
        if not os.path.exists(dir):
            os.makedirs(dir)

        with open(dest_file_name, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        return True, dest_file_name
    except Exception as e:
        print ("save_file %s , error:" % (dest_file_name), e)
        return False, None


    return True, dest_file_name

# 以文件形式存储内容
def save_as_file(content, file_name, dir = ""):
    dest_file_name = dir + '/' + file_name

    try:
        # 创建目录
        if not os.path.exists(dir):
            os.makedirs(dir)

        with open(dest_file_name, 'wb+') as destination:
            destination.write(content)
        return True, dest_file_name
    except Exception as e:
        print ("save_file %s , error:" % (dest_file_name), e)
        return False, None
    return True, dest_file_name


#file_name: /xxx/xxx/xxx  带绝对路径的文件
def remove_file(file_name):
    try:
        os.remove(file_name)
        return True
    except Exception as e:
        print("remove_file:%s err:" % (file_name), e)
        return False


'''
文件保存并最终删除的控制器
'''
class SaveDeleteFile:
    '''
    file 文件对象
    dir  本地的绝对文件路径
    file_name 保存的名字, 不带路径的
    '''
    def __init__(self, del_when_leave=True):
        self.saved = False
        self.file_name = None
        self.del_when_leave = del_when_leave

    def save_file(self,  file, file_name, dir = "UPLOAD_FILES_DIR"):
        self.saved, self.file_name = save_file(file=file, file_name=file_name, dir=dir)

    def save_as_file(self,  content, file_name, dir = "UPLOAD_FILES_DIR"):
        self.saved, self.file_name = save_as_file(content=content, file_name=file_name, dir=dir)

    def save_excel(self, file_name, sheet_name="qas", dir = "UPLOAD_FILES_DIR"):
        self.saved, self.file_name = save_excel(file_name=file_name, sheet_name=sheet_name, dir=dir)

    def save_ok(self):
        return self.saved == True and self.file_name is not None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.save_ok():
            if self.del_when_leave == True:
                remove_file(file_name=self.file_name)
            self.file_name = None
            self.saved = False





'''
针对excel的操作
'''
class WRExcel:
    '''
    file_name: /xxx/xxx/xxx.xlsx 带绝对路径的文件
    '''
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.excel_file = None
        self.excel_sheet = None


    def __enter__(self):
        #打开文件
        try:
            self.excel_file = openpyxl.load_workbook(self.file_name)
        except Exception as e:
            print("WRExcel open file:%s error:" % (self.file_name), e)
            self.excel_file = None
            return None
        #打开sheet
        try:
            self.excel_sheet = self.excel_file.get_sheet_by_name(self.sheet_name)
        except Exception as e:
            print("WRExcel open file:%s sheet:%s error:" % (self.file_name, self.sheet_name), e)
            self.excel_sheet = None
            self.excel_file.close()
            self.excel_file = None
            return None
        return self



    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.excel_file is not None:
            self.excel_file.close()

        self.excel_sheet = None
        self.excel_file = None

    #获取最大的行数
    def get_sheet_max_rows(self):
        if self.excel_sheet is not None:
            return self.excel_sheet.max_row
        return 0

    #获取最大的列数
    def get_sheet_max_columns(self):
        if self.excel_sheet is not None:
            return self.excel_sheet.max_column
        return 0

    def save(self):
        try:
            self.excel_file.save(filename=self.file_name)
            return True
        except Exception as e:
            print ('save excel %s error:' % (self.file_name), e)
            return False


# 获取第一个sheet
class FirstSheet(WRExcel):
    def __init__(self, file_name):
        self.file_name = file_name
        self.excel_file = None
        self.excel_sheet = None

    def __enter__(self):
        #打开文件
        try:
            self.excel_file = openpyxl.load_workbook(self.file_name)
        except Exception as e:
            print("WRExcel open file:%s error:" % (self.file_name), e)
            self.excel_file = None
            return None
        #打开sheet
        self.excel_file._active_sheet_index = 0
        self.excel_sheet = self.excel_file.active
        return self


#dir  本地的绝对文件路径
#file_name 保存的名字, 不带路径的
# sheet_name: 第一个sheet的名字
# return: True succ, False fail
def save_excel(file_name, sheet_name="qas", dir = "UPLOAD_FILES_DIR"):
    dest_file_name = dir + '/' + file_name

    # 新建一个workbook
    wb = Workbook()
    # 第一个sheet是ws
    ws = wb.worksheets[0]
    # 设置ws的名称
    ws.title = sheet_name
    try:
        wb.save(filename=dest_file_name)
        return True, dest_file_name
    except Exception as e:
        print ("save file(%s, %s) fail:" % (dest_file_name, sheet_name), e)
        return False, None




















